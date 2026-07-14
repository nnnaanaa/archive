"""
image フォルダ内の PNG 画像を Excel に貼り付けてエビデンスシートを生成する。

ファイル名規則: {グループ}_{画面名}.png
  - グループ: 先頭の固定文字列（例: 2-1, 2-2）
  - 画面名: 最後の _ より後の文字列（任意）
  - 分割は最後の _ で行うため、グループ名自体に _ が含まれていても問題ない

列の表示順は screen_order.txt に 1 行 1 画面名で記載する。
未記載の画面名はアルファベット順で末尾に追加される。
グループ 20 件ごとに 1 シートへ分割する。
"""

import datetime
from pathlib import Path
from collections import defaultdict

from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_points

IMAGE_DIR    = Path(__file__).parent / "image"
OUTPUT_FILE  = Path(__file__).parent / "evidence.xlsx"
ORDER_FILE   = Path(__file__).parent / "screen_order.txt"
GROUPS_PER_SHEET = 20

FONT_NAME   = "Meiryo UI"
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11)
LABEL_FONT  = Font(name=FONT_NAME, bold=True, size=11)
BODY_FONT   = Font(name=FONT_NAME, size=10)

PX_TO_COL = 8  # Excelの列幅1単位 ≒ 8px


def load_screen_order() -> list[str]:
    """screen_order.txt が存在すれば順序リストを返す。なければ空リスト。"""
    if not ORDER_FILE.exists():
        return []
    lines = ORDER_FILE.read_text(encoding="utf-8").splitlines()
    return [ln.strip() for ln in lines if ln.strip() and not ln.startswith("#")]


def sort_screens(screens: set[str], order: list[str]) -> list[str]:
    """指定順 → 未指定分をアルファベット順で後ろに並べる。"""
    ordered = [s for s in order if s in screens]
    rest    = sorted(screens - set(ordered))
    return ordered + rest


def collect_images(image_dir: Path) -> tuple[dict[str, dict[str, Path]], set[str]]:
    """最後の _ でグループと画面名に分割して収集する。"""
    groups: dict[str, dict[str, Path]] = defaultdict(dict)
    all_screens: set[str] = set()

    for f in sorted(image_dir.glob("*.png")):
        stem = f.stem
        if "_" in stem:
            idx    = stem.rfind("_")
            key    = stem[:idx]
            screen = stem[idx + 1:]
        else:
            key    = stem
            screen = "-"

        groups[key][screen] = f
        all_screens.add(screen)

    return dict(groups), all_screens


def image_size(path: Path) -> tuple[int, int]:
    with PILImage.open(path) as im:
        return im.size  # (width, height)


def add_image_to_cell(ws, img_path: Path, col: int, row: int):
    img = XLImage(str(img_path))
    ws.add_image(img, f"{get_column_letter(col)}{row}")


def write_sheet(wb: Workbook, sheet_name: str, keys: list[str],
                groups: dict[str, dict[str, Path]], screens: list[str],
                screen_max_w: dict[str, int], group_max_h: dict[str, int]):
    ws = wb.create_sheet(title=sheet_name)

    # 列幅設定
    ws.column_dimensions["A"].width = 15
    for i, screen in enumerate(screens, start=2):
        ws.column_dimensions[get_column_letter(i)].width = screen_max_w[screen] / PX_TO_COL + 2

    current_row = 1

    # ===== 画面名ヘッダー行 =====
    c = ws.cell(row=current_row, column=1, value="番号")
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    for i, screen in enumerate(screens, start=2):
        c = ws.cell(row=current_row, column=i, value=screen)
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # ===== 各グループの行 =====
    for key in keys:
        group = groups[key]

        # 画像行
        lc = ws.cell(row=current_row, column=1, value=key)
        lc.font = LABEL_FONT
        lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = pixels_to_points(group_max_h.get(key, 300))

        for i, screen in enumerate(screens, start=2):
            if screen in group:
                add_image_to_cell(ws, group[screen], col=i, row=current_row)

        current_row += 1

        # ファイル名行
        ws.row_dimensions[current_row].height = 18
        for i, screen in enumerate(screens, start=2):
            if screen in group:
                c = ws.cell(row=current_row, column=i, value=group[screen].name)
                c.font = BODY_FONT
                c.alignment = Alignment(horizontal="center", vertical="center")

        current_row += 1


def build_excel(groups: dict[str, dict[str, Path]], screens: list[str], output: Path):
    # 各画面の最大画像幅・グループごとの最大高さを事前計算
    screen_max_w: dict[str, int] = {s: 0 for s in screens}
    group_max_h:  dict[str, int] = {}
    for key, group in groups.items():
        max_h = 0
        for screen, path in group.items():
            w, h = image_size(path)
            if screen in screen_max_w:
                screen_max_w[screen] = max(screen_max_w[screen], w)
            max_h = max(max_h, h)
        group_max_h[key] = max_h

    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    sorted_keys = sorted(groups.keys())
    total = len(sorted_keys)
    num_sheets = (total + GROUPS_PER_SHEET - 1) // GROUPS_PER_SHEET

    for page in range(num_sheets):
        start = page * GROUPS_PER_SHEET
        end   = min(start + GROUPS_PER_SHEET, total)
        keys_chunk = sorted_keys[start:end]
        sheet_name = f"エビデンス({page + 1})" if num_sheets > 1 else "エビデンス一覧"
        write_sheet(wb, sheet_name, keys_chunk, groups, screens, screen_max_w, group_max_h)

    # ===== ファイル一覧シート =====
    ws2 = wb.create_sheet("ファイル一覧")
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 22

    for col, h in enumerate(["番号", "ファイル名", "画面名", "更新日時"], start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 20

    r = 2
    for key in sorted_keys:
        for screen in screens:
            if screen in groups[key]:
                p = groups[key][screen]
                mtime = datetime.datetime.fromtimestamp(p.stat().st_mtime)
                c1 = ws2.cell(row=r, column=1, value=key)
                c1.font = BODY_FONT; c1.alignment = Alignment(horizontal="center")
                c2 = ws2.cell(row=r, column=2, value=p.name)
                c2.font = BODY_FONT
                c3 = ws2.cell(row=r, column=3, value=screen)
                c3.font = BODY_FONT; c3.alignment = Alignment(horizontal="center")
                c4 = ws2.cell(row=r, column=4, value=mtime.strftime("%Y/%m/%d %H:%M:%S"))
                c4.font = BODY_FONT; c4.alignment = Alignment(horizontal="center")
                r += 1

    wb.save(output)
    print(f"保存完了: {output}")


def main():
    if not IMAGE_DIR.exists():
        raise FileNotFoundError(f"画像フォルダが見つかりません: {IMAGE_DIR}")

    groups, all_screens = collect_images(IMAGE_DIR)
    if not groups:
        print("対象画像が見つかりませんでした。")
        return

    order   = load_screen_order()
    screens = sort_screens(all_screens, order)

    print(f"検出グループ: {sorted(groups.keys())}")
    print(f"画面順:       {screens}")
    if order:
        print(f"  (screen_order.txt を使用)")
    else:
        print(f"  (アルファベット順 / screen_order.txt がなければこの順)")

    build_excel(groups, screens, OUTPUT_FILE)


if __name__ == "__main__":
    main()
