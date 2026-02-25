import os
import glob
import math
import re
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

def import_bats_to_excel_smart_analysis(output_name="bat_analysis_document.xlsx"):
    wb = Workbook()
    first_sheet = True

    # スタイル設定
    side = Side(style='thin', color='000000')
    font_meiryo = Font(name='Meiryo UI', size=11, bold=False)
    font_note = Font(name='Meiryo UI', size=10, bold=False)
    
    fills = {
        "red": PatternFill(start_color="FFD1D1", end_color="FFD1D1", fill_type="solid"),
        "blue": PatternFill(start_color="D1EAFF", end_color="D1EAFF", fill_type="solid"),
        "green": PatternFill(start_color="D1FFD1", end_color="D1FFD1", fill_type="solid"),
        "yellow": PatternFill(start_color="FFF9D1", end_color="FFF9D1", fill_type="solid"),
        "purple": PatternFill(start_color="EAD1FF", end_color="EAD1FF", fill_type="solid"), # 制御構文用
    }

    # 基本キーワード
    keyword_map = {
        "DEL ": (fills["red"], "ファイル削除"), "ERASE ": (fills["red"], "ファイル削除"),
        "RMDIR ": (fills["red"], "フォルダ削除"), "RD ": (fills["red"], "フォルダ削除"),
        "COPY ": (fills["blue"], "コピー"), "XCOPY ": (fills["blue"], "ディレクトリコピー"),
        "ROBOCOPY ": (fills["blue"], "高度な同期"), "MOVE ": (fills["blue"], "移動"),
        "MKDIR ": (fills["green"], "フォルダ作成"), "MD ": (fills["green"], "フォルダ作成"),
        "NET USE": (fills["green"], "NW接続"),
        "PAUSE": (fills["yellow"], "一時停止"),
    }

    # 制御構文の解析関数
    def analyze_control_syntax(line):
        upper = line.upper().strip()
        # IF文の解析
        if upper.startswith("IF "):
            if "EXIST " in upper: return fills["purple"], "条件分岐：ファイルの存在確認"
            if "DEFINED " in upper: return fills["purple"], "条件分岐：変数の定義確認"
            if "==" in upper or " EQU " in upper: return fills["purple"], "条件分岐：値の比較"
            if "ERRORLEVEL" in upper: return fills["yellow"], "条件分岐：エラー判定"
            return fills["purple"], "条件分岐(IF)"
        
        # FOR文の解析
        if upper.startswith("FOR "):
            if "/F" in upper: return fills["purple"], "ループ：ファイル内やコマンド結果の解析"
            if "/R" in upper: return fills["purple"], "ループ：サブフォルダまで再帰処理"
            if "/D" in upper: return fills["purple"], "ループ：ディレクトリを対象に処理"
            if "/L" in upper: return fills["purple"], "ループ：数値カウント（開始,ステップ,終了）"
            return fills["purple"], "ループ処理(FOR)"
        
        return None, ""

    target_dir = os.path.join(os.getcwd(), "bat")
    if not os.path.exists(target_dir):
        print(f"エラー: '{target_dir}' フォルダが見算かりません。")
        return

    bat_files = glob.glob(os.path.join(target_dir, "*.bat"))

    for file_path in bat_files:
        file_name = os.path.basename(file_path)
        sheet_title = file_name[:31]
        ws = wb.active if first_sheet else wb.create_sheet(title=sheet_title)
        if first_sheet: ws.title = sheet_title; first_sheet = False

        try:
            with open(file_path, "r", encoding="shift_jis") as f:
                lines = [line.rstrip() for line in f.readlines()]
        except:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                lines = [line.rstrip() for line in f.readlines()]

        if not lines: lines = ["(空のファイル)"]

        # 最大列数の計算
        max_bytes = max(len(t.encode('shift_jis', errors='ignore')) for t in [file_name] + lines)
        max_col = math.ceil(max_bytes / 2.0) + 2

        for i in range(1, max_col + 15): # 補足が長くなる可能性があるので少し多めに
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 2.5

        # タイトル行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = file_name
        title_cell.font = font_meiryo
        for c in range(1, max_col + 1):
            ws.cell(row=1, column=c).border = Border(top=side, bottom=side, left=(side if c==1 else None), right=(side if c==max_col else None))

        # ソース行
        last_row = len(lines) + 1
        for r_idx, content in enumerate(lines, start=2):
            # 1. 基本キーワードチェック
            target_fill = None
            note_text = ""
            upper_content = content.upper()
            
            for kw, (fill, note) in keyword_map.items():
                if kw in upper_content:
                    target_fill, note_text = fill, note
                    break
            
            # 2. 制御構文(IF/FOR)の解析（基本キーワードより優先）
            syntax_fill, syntax_note = analyze_control_syntax(content)
            if syntax_fill:
                target_fill, note_text = syntax_fill, syntax_note

            # セル描画
            for c_idx in range(1, max_col + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if c_idx == 1:
                    cell.value = content
                    cell.font = font_meiryo
                if target_fill: cell.fill = target_fill
                cell.border = Border(top=(side if r_idx == 2 else None), bottom=(side if r_idx == last_row else None), 
                                     left=(side if c_idx == 1 else None), right=(side if c_idx == max_col else None))

            # 補足
            if note_text:
                note_cell = ws.cell(row=r_idx, column=max_col + 1)
                note_cell.value = "← " + note_text
                note_cell.font = font_note

        print(f"解析完了: {file_name}")

    wb.save(output_name)
    print(f"\nすべての処理が完了しました。ファイル名: {output_name}")

if __name__ == "__main__":
    import_bats_to_excel_smart_analysis()