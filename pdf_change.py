import pdfplumber
import openpyxl
import re

def clean_text(text):
    """セル内の不要な空白や改行を整理する"""
    if text is None:
        return ""
    # 改行や複数スペースを整理
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def convert_pdf_to_excel_sheets(pdf_path, output_path):
    # Excelワークブックを作成
    wb = openpyxl.Workbook()
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages):
            # ページごとにシートを作成
            ws = wb.create_sheet(title=f"ページ{page_no + 1}")
            print(f"処理中: ページ{page_no + 1}...")
            
            # 表を抽出
            table = page.extract_table({"vertical_strategy": "lines", "horizontal_strategy": "lines"})
            
            if table:
                for row_idx, row in enumerate(table):
                    for col_idx, cell in enumerate(row):
                        ws.cell(row=row_idx + 1, column=col_idx + 1, value=clean_text(cell))
            else:
                ws.cell(row=1, column=1, value="このページには抽出可能な表データがありませんでした")
    
    # 最初のデフォルトシート（Sheet）が空なら削除
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])
    
    # 保存
    wb.save(output_path)
    print(f"完了: {output_path} を保存しました。")

# 実行
convert_pdf_to_excel_sheets("your_file.pdf", "output_pages.xlsx")